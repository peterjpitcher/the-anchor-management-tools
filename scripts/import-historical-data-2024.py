#!/usr/bin/env python3
"""
Import historical 2024 data from Excel spreadsheets into Supabase SQL migration.

Reads:
  1. Mileage trips (Jan-Feb 2024) from Mileage Tracker spreadsheet
  2. Expenses (Jan-Dec 2024) from Petty Cash spreadsheet

Outputs:
  supabase/migrations/20260405130004_import_historical_data_2024.sql
"""

import os
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
MIGRATION_FILE = PROJECT_ROOT / "supabase" / "migrations" / "20260405130004_import_historical_data_2024.sql"

ICLOUD = Path.home() / "Library" / "Mobile Documents" / "com~apple~CloudDocs"
FINANCE_DIR = ICLOUD / "1. The Anchor" / "Financials and Accounting"

MILEAGE_FILE = FINANCE_DIR / "Mileage I Petty Cash 2024" / "Mileage Tracker - 2024.xlsx"
EXPENSES_FILE = FINANCE_DIR / "Mileage I Petty Cash 2024" / "Petty Cash - 2024.xlsx"

HMRC_STANDARD_RATE = 0.45
MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sql_date(val) -> str:
    """Convert a datetime or date to 'YYYY-MM-DD' string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    raise ValueError(f"Cannot convert {val!r} to date")


def sql_text(val) -> str:
    """Escape a string for SQL single-quoted literals."""
    if val is None:
        return "NULL"
    s = str(val).strip()
    if not s:
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def sql_num(val, decimals=2) -> str:
    """Format a number for SQL."""
    if val is None:
        return "0"
    return f"{float(val):.{decimals}f}"


def sql_bool(val) -> str:
    """Convert to SQL boolean."""
    if val is None:
        return "FALSE"
    if isinstance(val, str):
        return "TRUE" if val.strip().upper() in ("YES", "TRUE", "1") else "FALSE"
    return "TRUE" if val else "FALSE"


# ---------------------------------------------------------------------------
# 1. Mileage
# ---------------------------------------------------------------------------

def read_mileage_trips():
    """Read mileage trips from Jan and Feb sheets."""
    wb = openpyxl.load_workbook(str(MILEAGE_FILE), data_only=True)
    trips = []

    # 2024 mileage only has Jan-Feb data (37 trips)
    for month in ["January", "February"]:
        if month not in wb.sheetnames:
            print(f"  WARNING: Sheet '{month}' not found in mileage workbook, skipping")
            continue
        ws = wb[month]
        for row in range(4, 100):
            date_val = ws.cell(row=row, column=1).value
            dest = ws.cell(row=row, column=2).value
            miles = ws.cell(row=row, column=4).value

            if date_val is None or dest is None or miles is None:
                continue

            miles = float(miles)
            if miles <= 0:
                continue

            # Fix year typos: data is in 2024 sheets, so 2025 dates are typos
            if isinstance(date_val, datetime) and date_val.year == 2025:
                date_val = date_val.replace(year=2024)

            trips.append({
                "trip_date": date_val,
                "destination": str(dest).strip(),
                "total_miles": miles,  # This is RETURN miles
            })

    print(f"  Mileage: {len(trips)} trips read")
    return trips


def generate_mileage_sql(trips):
    """Generate SQL for mileage trips and legs."""
    lines = []
    lines.append("-- =============================================================")
    lines.append("-- Mileage Trips (historical import: Jan-Feb 2024)")
    lines.append("-- =============================================================")
    lines.append("")

    # Collect unique destinations from trips
    dest_names = sorted(set(t["destination"] for t in trips))

    # Insert any destinations not already in the DB
    lines.append("-- Insert any destinations not already seeded")
    for name in dest_names:
        lines.append(
            f"INSERT INTO public.mileage_destinations (name) "
            f"SELECT {sql_text(name)} WHERE NOT EXISTS ("
            f"SELECT 1 FROM public.mileage_destinations WHERE name = {sql_text(name)}"
            f");"
        )
    lines.append("")

    # Insert trips using a DO block so we can capture trip IDs for legs
    lines.append("-- Insert trips and their legs")
    lines.append("DO $$")
    lines.append("DECLARE")
    lines.append("  v_trip_id UUID;")
    lines.append("  v_anchor_id UUID;")
    lines.append("  v_dest_id UUID;")
    lines.append("BEGIN")
    lines.append("  -- Get the Anchor (home base) destination ID")
    lines.append("  SELECT id INTO v_anchor_id FROM public.mileage_destinations WHERE is_home_base = TRUE LIMIT 1;")
    lines.append("")

    for i, trip in enumerate(trips):
        trip_date = sql_date(trip["trip_date"])
        dest_name = trip["destination"]
        total_miles = trip["total_miles"]
        half_miles = round(total_miles / 2, 1)
        amount_due = round(total_miles * HMRC_STANDARD_RATE, 2)

        lines.append(f"  -- Trip {i+1}: {trip_date} -> {dest_name} ({total_miles} miles return)")
        lines.append(f"  SELECT id INTO v_dest_id FROM public.mileage_destinations WHERE name = {sql_text(dest_name)} LIMIT 1;")
        lines.append(f"  INSERT INTO public.mileage_trips (trip_date, description, total_miles, miles_at_standard_rate, miles_at_reduced_rate, amount_due, source, created_by)")
        lines.append(f"  VALUES ('{trip_date}', {sql_text(dest_name)}, {sql_num(total_miles, 1)}, {sql_num(total_miles, 1)}, 0, {sql_num(amount_due, 2)}, 'manual', NULL)")
        lines.append(f"  RETURNING id INTO v_trip_id;")
        lines.append("")

        # Leg 1: Anchor -> Destination
        lines.append(f"  INSERT INTO public.mileage_trip_legs (trip_id, leg_order, from_destination_id, to_destination_id, miles)")
        lines.append(f"  VALUES (v_trip_id, 1, v_anchor_id, v_dest_id, {sql_num(half_miles, 1)});")

        # Leg 2: Destination -> Anchor
        lines.append(f"  INSERT INTO public.mileage_trip_legs (trip_id, leg_order, from_destination_id, to_destination_id, miles)")
        lines.append(f"  VALUES (v_trip_id, 2, v_dest_id, v_anchor_id, {sql_num(half_miles, 1)});")
        lines.append("")

    lines.append("END;")
    lines.append("$$;")
    lines.append("")

    return lines


# ---------------------------------------------------------------------------
# 2. Expenses
# ---------------------------------------------------------------------------

def read_expenses():
    """Read expenses from all monthly sheets."""
    wb = openpyxl.load_workbook(str(EXPENSES_FILE), data_only=True)
    expenses = []

    for month in MONTHS:
        if month not in wb.sheetnames:
            print(f"  WARNING: Sheet '{month}' not found in expenses workbook, skipping")
            continue
        ws = wb[month]
        # Header is row 4, data starts row 5
        for row in range(5, 200):
            date_val = ws.cell(row=row, column=1).value
            company_ref = ws.cell(row=row, column=2).value
            justification = ws.cell(row=row, column=3).value
            amount = ws.cell(row=row, column=4).value
            vat_applicable = ws.cell(row=row, column=5).value
            vat_amount = ws.cell(row=row, column=6).value

            if date_val is None or amount is None:
                continue

            # Skip non-numeric amounts (e.g. separator rows with '***')
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                continue
            if amount <= 0:
                continue

            # Fix year typos: data is in 2024 sheets, so 2025 dates are typos
            if isinstance(date_val, datetime) and date_val.year == 2025:
                date_val = date_val.replace(year=2024)

            # Default justification to company_ref if NULL
            just_str = str(justification).strip() if justification else None
            comp_str = str(company_ref).strip() if company_ref else "Unknown"
            if not just_str:
                just_str = comp_str

            expenses.append({
                "expense_date": date_val,
                "company_ref": comp_str,
                "justification": just_str,
                "amount": amount,
                "vat_applicable": vat_applicable,
                "vat_amount": float(vat_amount) if vat_amount else 0,
            })

    print(f"  Expenses: {len(expenses)} entries read")
    return expenses


def generate_expenses_sql(expenses):
    """Generate SQL for expenses."""
    lines = []
    lines.append("-- =============================================================")
    lines.append("-- Expenses (historical import: Jan-Dec 2024)")
    lines.append("-- =============================================================")
    lines.append("")
    lines.append("INSERT INTO public.expenses (expense_date, company_ref, justification, amount, vat_applicable, vat_amount, created_by) VALUES")

    value_lines = []
    for exp in expenses:
        vat_app = sql_bool(exp["vat_applicable"])
        vat_amt = max(0, exp["vat_amount"])  # Ensure non-negative
        value_lines.append(
            f"  ('{sql_date(exp['expense_date'])}', {sql_text(exp['company_ref'])}, "
            f"{sql_text(exp['justification'])}, {sql_num(exp['amount'])}, "
            f"{vat_app}, {sql_num(vat_amt)}, NULL)"
        )

    lines.append(",\n".join(value_lines) + ";")
    lines.append("")

    return lines


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Reading 2024 spreadsheets...")

    # Verify files exist
    for name, path in [("Mileage", MILEAGE_FILE), ("Expenses", EXPENSES_FILE)]:
        if not path.exists():
            print(f"ERROR: {name} file not found: {path}")
            sys.exit(1)

    # Read data
    mileage_trips = read_mileage_trips()
    expenses = read_expenses()

    # Generate SQL
    print("\nGenerating SQL migration...")
    sql_lines = []
    sql_lines.append("-- =============================================================")
    sql_lines.append("-- Migration: Import Historical 2024 Data from Excel Spreadsheets")
    sql_lines.append("-- Generated by: scripts/import-historical-data-2024.py")
    sql_lines.append(f"-- Generated at: {datetime.now().isoformat()}")
    sql_lines.append("-- =============================================================")
    sql_lines.append("-- Data sources:")
    sql_lines.append("--   1. Mileage Tracker - 2024.xlsx (Jan-Feb, 37 trips)")
    sql_lines.append("--   2. Petty Cash - 2024.xlsx (Jan-Dec, 335 expenses)")
    sql_lines.append("-- =============================================================")
    sql_lines.append("")

    sql_lines.extend(generate_mileage_sql(mileage_trips))
    sql_lines.extend(generate_expenses_sql(expenses))
    sql_lines.append("")

    # Write migration file
    MIGRATION_FILE.parent.mkdir(parents=True, exist_ok=True)
    MIGRATION_FILE.write_text("\n".join(sql_lines), encoding="utf-8")

    print(f"\nMigration written to: {MIGRATION_FILE}")
    print(f"  Mileage trips: {len(mileage_trips)}")
    print(f"  Expenses: {len(expenses)}")
    print("\nTo apply: npx supabase db push --include-all")


if __name__ == "__main__":
    main()
