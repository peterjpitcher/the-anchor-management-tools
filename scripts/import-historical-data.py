#!/usr/bin/env python3
"""
Import historical data from Excel spreadsheets into Supabase SQL migration.

Reads:
  1. Mileage trips (Oct-Dec 2025) from Mileage Tracker spreadsheet
  2. Expenses (Jan-Dec 2025) from Petty Cash spreadsheet
  3. MGD collections + returns from Machine Games Duty spreadsheet

Outputs:
  supabase/migrations/20260405130003_import_historical_data.sql
"""

import os
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
MIGRATION_FILE = PROJECT_ROOT / "supabase" / "migrations" / "20260405130003_import_historical_data.sql"

ICLOUD = Path.home() / "Library" / "Mobile Documents" / "com~apple~CloudDocs"
FINANCE_DIR = ICLOUD / "1. The Anchor" / "Financials and Accounting"

MILEAGE_FILE = FINANCE_DIR / "Mileage I Petty Cash 2024" / "Mileage Tracker - 2025.xlsx"
EXPENSES_FILE = FINANCE_DIR / "Mileage I Petty Cash 2024" / "Petty Cash - 2025.xlsx"
MGD_FILE = FINANCE_DIR / "MGD" / "Machine Games Duty Collections (New).xlsm"

HMRC_STANDARD_RATE = 0.45
MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sql_date(val) -> str:
    """Convert a datetime or date to 'YYYY-MM-DD' string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    raise ValueError(f"Cannot convert {val!r} to date")


def sql_text(val) -> str:
    """Escape a string for SQL single-quoted literals."""
    if val is None:
        return "NULL"
    s = str(val).strip()
    if not s:
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def sql_num(val, decimals=2) -> str:
    """Format a number for SQL."""
    if val is None:
        return "0"
    return f"{float(val):.{decimals}f}"


def sql_bool(val) -> str:
    """Convert to SQL boolean."""
    if val is None:
        return "FALSE"
    if isinstance(val, str):
        return "TRUE" if val.strip().upper() in ("YES", "TRUE", "1") else "FALSE"
    return "TRUE" if val else "FALSE"


# ---------------------------------------------------------------------------
# 1. Mileage
# ---------------------------------------------------------------------------

def read_mileage_trips():
    """Read mileage trips from Oct, Nov, Dec sheets."""
    wb = openpyxl.load_workbook(str(MILEAGE_FILE), data_only=True)
    trips = []

    for month in ["October", "November", "December"]:
        ws = wb[month]
        for row in range(4, 100):
            date_val = ws.cell(row=row, column=1).value
            dest = ws.cell(row=row, column=2).value
            miles = ws.cell(row=row, column=4).value

            if date_val is None or dest is None or miles is None:
                continue

            miles = float(miles)
            if miles <= 0:
                continue

            # Fix year typos: data is in 2025 sheets, so 2024 dates are typos
            if isinstance(date_val, datetime) and date_val.year == 2024:
                date_val = date_val.replace(year=2025)

            trips.append({
                "trip_date": date_val,
                "destination": str(dest).strip(),
                "total_miles": miles,  # This is RETURN miles
            })

    print(f"  Mileage: {len(trips)} trips read")
    return trips


def generate_mileage_sql(trips):
    """Generate SQL for mileage trips and legs."""
    lines = []
    lines.append("-- =============================================================")
    lines.append("-- Mileage Trips (historical import: Oct-Dec 2025)")
    lines.append("-- =============================================================")
    lines.append("")

    # Collect unique destinations from trips
    dest_names = sorted(set(t["destination"] for t in trips))

    # Check which destinations might be missing from seeded data
    # (The migration already seeds 43 destinations)
    lines.append("-- Insert any destinations not already seeded")
    for name in dest_names:
        lines.append(
            f"INSERT INTO public.mileage_destinations (name) "
            f"SELECT {sql_text(name)} WHERE NOT EXISTS ("
            f"SELECT 1 FROM public.mileage_destinations WHERE name = {sql_text(name)}"
            f");"
        )
    lines.append("")

    # Insert trips using a DO block so we can capture trip IDs for legs
    lines.append("-- Insert trips and their legs")
    lines.append("DO $$")
    lines.append("DECLARE")
    lines.append("  v_trip_id UUID;")
    lines.append("  v_anchor_id UUID;")
    lines.append("  v_dest_id UUID;")
    lines.append("BEGIN")
    lines.append("  -- Get the Anchor (home base) destination ID")
    lines.append("  SELECT id INTO v_anchor_id FROM public.mileage_destinations WHERE is_home_base = TRUE LIMIT 1;")
    lines.append("")

    for i, trip in enumerate(trips):
        trip_date = sql_date(trip["trip_date"])
        dest_name = trip["destination"]
        total_miles = trip["total_miles"]
        half_miles = round(total_miles / 2, 1)
        amount_due = round(total_miles * HMRC_STANDARD_RATE, 2)

        lines.append(f"  -- Trip {i+1}: {trip_date} -> {dest_name} ({total_miles} miles return)")
        lines.append(f"  SELECT id INTO v_dest_id FROM public.mileage_destinations WHERE name = {sql_text(dest_name)} LIMIT 1;")
        lines.append(f"  INSERT INTO public.mileage_trips (trip_date, description, total_miles, miles_at_standard_rate, miles_at_reduced_rate, amount_due, source, created_by)")
        lines.append(f"  VALUES ('{trip_date}', {sql_text(dest_name)}, {sql_num(total_miles, 1)}, {sql_num(total_miles, 1)}, 0, {sql_num(amount_due, 2)}, 'manual', NULL)")
        lines.append(f"  RETURNING id INTO v_trip_id;")
        lines.append("")

        # Leg 1: Anchor -> Destination
        lines.append(f"  INSERT INTO public.mileage_trip_legs (trip_id, leg_order, from_destination_id, to_destination_id, miles)")
        lines.append(f"  VALUES (v_trip_id, 1, v_anchor_id, v_dest_id, {sql_num(half_miles, 1)});")

        # Leg 2: Destination -> Anchor
        lines.append(f"  INSERT INTO public.mileage_trip_legs (trip_id, leg_order, from_destination_id, to_destination_id, miles)")
        lines.append(f"  VALUES (v_trip_id, 2, v_dest_id, v_anchor_id, {sql_num(half_miles, 1)});")
        lines.append("")

    lines.append("END;")
    lines.append("$$;")
    lines.append("")

    return lines


# ---------------------------------------------------------------------------
# 2. Expenses
# ---------------------------------------------------------------------------

def read_expenses():
    """Read expenses from all monthly sheets."""
    wb = openpyxl.load_workbook(str(EXPENSES_FILE), data_only=True)
    expenses = []

    for month in MONTHS:
        ws = wb[month]
        # Header is row 4, data starts row 5
        for row in range(5, 200):
            date_val = ws.cell(row=row, column=1).value
            company_ref = ws.cell(row=row, column=2).value
            justification = ws.cell(row=row, column=3).value
            amount = ws.cell(row=row, column=4).value
            vat_applicable = ws.cell(row=row, column=5).value
            vat_amount = ws.cell(row=row, column=6).value

            if date_val is None or amount is None:
                continue

            amount = float(amount)
            if amount <= 0:
                continue

            # Fix year typos: data is in 2025 sheets, so 2024 dates are typos
            if isinstance(date_val, datetime) and date_val.year == 2024:
                date_val = date_val.replace(year=2025)

            expenses.append({
                "expense_date": date_val,
                "company_ref": str(company_ref).strip() if company_ref else "Unknown",
                "justification": str(justification).strip() if justification else "No justification",
                "amount": amount,
                "vat_applicable": vat_applicable,
                "vat_amount": float(vat_amount) if vat_amount else 0,
            })

    print(f"  Expenses: {len(expenses)} entries read")
    return expenses


def generate_expenses_sql(expenses):
    """Generate SQL for expenses."""
    lines = []
    lines.append("-- =============================================================")
    lines.append("-- Expenses (historical import: Jan-Dec 2025)")
    lines.append("-- =============================================================")
    lines.append("")
    lines.append("INSERT INTO public.expenses (expense_date, company_ref, justification, amount, vat_applicable, vat_amount, created_by) VALUES")

    value_lines = []
    for exp in expenses:
        vat_app = sql_bool(exp["vat_applicable"])
        vat_amt = max(0, exp["vat_amount"])  # Ensure non-negative
        value_lines.append(
            f"  ('{sql_date(exp['expense_date'])}', {sql_text(exp['company_ref'])}, "
            f"{sql_text(exp['justification'])}, {sql_num(exp['amount'])}, "
            f"{vat_app}, {sql_num(vat_amt)}, NULL)"
        )

    lines.append(",\n".join(value_lines) + ";")
    lines.append("")

    return lines


# ---------------------------------------------------------------------------
# 3. MGD
# ---------------------------------------------------------------------------

def read_mgd_collections():
    """Read MGD collections from left side of Takings sheet."""
    wb = openpyxl.load_workbook(str(MGD_FILE), data_only=True)
    ws = wb["Takings"]
    collections = []

    for row in range(4, 200):
        date_val = ws.cell(row=row, column=1).value
        net_take = ws.cell(row=row, column=2).value
        # Column 3 is MGD (generated, skip)
        vat_on_supplier = ws.cell(row=row, column=4).value

        if date_val is None or net_take is None:
            continue

        net_take = float(net_take)
        vat_on_supplier = float(vat_on_supplier) if vat_on_supplier else 0

        collections.append({
            "collection_date": date_val,
            "net_take": net_take,
            "vat_on_supplier": max(0, vat_on_supplier),
        })

    print(f"  MGD Collections: {len(collections)} entries read")
    return collections


def read_mgd_returns():
    """Read MGD returns from right side of Takings sheet."""
    wb = openpyxl.load_workbook(str(MGD_FILE), data_only=True)
    ws = wb["Takings"]
    returns = []

    for row in range(4, 50):
        open_date = ws.cell(row=row, column=6).value
        close_date = ws.cell(row=row, column=7).value
        total_net = ws.cell(row=row, column=8).value
        total_mgd = ws.cell(row=row, column=9).value
        date_paid = ws.cell(row=row, column=10).value

        if open_date is None or close_date is None:
            continue

        total_net = float(total_net) if total_net else 0
        total_mgd = float(total_mgd) if total_mgd else 0

        # Normalise close_date to date object for comparison
        if isinstance(close_date, datetime):
            close_date_d = close_date.date()
        else:
            close_date_d = close_date

        # Determine status
        if date_paid is not None:
            status = "paid"
        elif close_date_d < date.today():
            status = "submitted"
        else:
            status = "open"

        # Skip future empty periods (net_take=0 and no collections expected)
        # Keep them - they were explicitly created in the spreadsheet

        returns.append({
            "period_start": open_date,
            "period_end": close_date,
            "total_net_take": total_net,
            "total_mgd": total_mgd,
            "status": status,
            "date_paid": date_paid,
        })

    print(f"  MGD Returns: {len(returns)} entries read")
    return returns


def generate_mgd_sql(collections, returns):
    """Generate SQL for MGD data."""
    lines = []
    lines.append("-- =============================================================")
    lines.append("-- MGD Returns (historical import)")
    lines.append("-- =============================================================")
    lines.append("-- Insert returns FIRST, then disable the collection trigger,")
    lines.append("-- insert collections, recalculate totals, then re-enable trigger.")
    lines.append("")

    # Disable the trigger so bulk insert doesn't fire it per row
    lines.append("-- Disable the auto-sync trigger during bulk import")
    lines.append("ALTER TABLE public.mgd_collections DISABLE TRIGGER trg_mgd_collection_sync;")
    lines.append("")

    # Insert returns
    lines.append("INSERT INTO public.mgd_returns (period_start, period_end, total_net_take, total_mgd, total_vat_on_supplier, status, date_paid) VALUES")
    value_lines = []
    for ret in returns:
        date_paid_sql = f"'{sql_date(ret['date_paid'])}'" if ret["date_paid"] else "NULL"
        # total_vat_on_supplier not in spreadsheet; will be recalculated from collections
        value_lines.append(
            f"  ('{sql_date(ret['period_start'])}', '{sql_date(ret['period_end'])}', "
            f"{sql_num(ret['total_net_take'])}, {sql_num(ret['total_mgd'])}, "
            f"0, '{ret['status']}', {date_paid_sql})"
        )
    lines.append(",\n".join(value_lines) + ";")
    lines.append("")

    # Insert collections (NOT including mgd_amount — it's GENERATED)
    lines.append("-- =============================================================")
    lines.append("-- MGD Collections (historical import)")
    lines.append("-- =============================================================")
    lines.append("")
    lines.append("INSERT INTO public.mgd_collections (collection_date, net_take, vat_on_supplier, created_by) VALUES")
    value_lines = []
    for coll in collections:
        value_lines.append(
            f"  ('{sql_date(coll['collection_date'])}', {sql_num(coll['net_take'])}, "
            f"{sql_num(coll['vat_on_supplier'])}, NULL)"
        )
    lines.append(",\n".join(value_lines) + ";")
    lines.append("")

    # Recalculate totals for all returns based on actual collection data
    lines.append("-- Recalculate return totals from actual collection data")
    lines.append("UPDATE public.mgd_returns r SET")
    lines.append("  total_net_take = COALESCE(sub.sum_net, 0),")
    lines.append("  total_mgd = COALESCE(sub.sum_mgd, 0),")
    lines.append("  total_vat_on_supplier = COALESCE(sub.sum_vat, 0)")
    lines.append("FROM (")
    lines.append("  SELECT")
    lines.append("    ret.id,")
    lines.append("    SUM(c.net_take) AS sum_net,")
    lines.append("    SUM(c.mgd_amount) AS sum_mgd,")
    lines.append("    SUM(c.vat_on_supplier) AS sum_vat")
    lines.append("  FROM public.mgd_returns ret")
    lines.append("  LEFT JOIN public.mgd_collections c")
    lines.append("    ON c.collection_date >= ret.period_start")
    lines.append("    AND c.collection_date <= ret.period_end")
    lines.append("  GROUP BY ret.id")
    lines.append(") sub")
    lines.append("WHERE r.id = sub.id;")
    lines.append("")

    # Re-enable the trigger
    lines.append("-- Re-enable the auto-sync trigger")
    lines.append("ALTER TABLE public.mgd_collections ENABLE TRIGGER trg_mgd_collection_sync;")
    lines.append("")

    return lines


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Reading spreadsheets...")

    # Verify files exist
    for name, path in [("Mileage", MILEAGE_FILE), ("Expenses", EXPENSES_FILE), ("MGD", MGD_FILE)]:
        if not path.exists():
            print(f"ERROR: {name} file not found: {path}")
            sys.exit(1)

    # Read data
    mileage_trips = read_mileage_trips()
    expenses = read_expenses()
    mgd_collections = read_mgd_collections()
    mgd_returns = read_mgd_returns()

    # Generate SQL
    print("\nGenerating SQL migration...")
    sql_lines = []
    sql_lines.append("-- =============================================================")
    sql_lines.append("-- Migration: Import Historical Data from Excel Spreadsheets")
    sql_lines.append("-- Generated by: scripts/import-historical-data.py")
    sql_lines.append(f"-- Generated at: {datetime.now().isoformat()}")
    sql_lines.append("-- =============================================================")
    sql_lines.append("-- Data sources:")
    sql_lines.append("--   1. Mileage Tracker - 2025.xlsx (Oct-Dec, 68 trips)")
    sql_lines.append("--   2. Petty Cash - 2025.xlsx (Jan-Dec, expenses)")
    sql_lines.append("--   3. Machine Games Duty Collections (New).xlsm (collections + returns)")
    sql_lines.append("-- =============================================================")
    sql_lines.append("")

    sql_lines.extend(generate_mileage_sql(mileage_trips))
    sql_lines.extend(generate_expenses_sql(expenses))
    sql_lines.extend(generate_mgd_sql(mgd_collections, mgd_returns))
    sql_lines.append("")

    # Write migration file
    MIGRATION_FILE.parent.mkdir(parents=True, exist_ok=True)
    MIGRATION_FILE.write_text("\n".join(sql_lines), encoding="utf-8")

    print(f"\nMigration written to: {MIGRATION_FILE}")
    print(f"  Mileage trips: {len(mileage_trips)}")
    print(f"  Expenses: {len(expenses)}")
    print(f"  MGD collections: {len(mgd_collections)}")
    print(f"  MGD returns: {len(mgd_returns)}")
    print("\nTo apply: npx supabase db push --include-all")


if __name__ == "__main__":
    main()
